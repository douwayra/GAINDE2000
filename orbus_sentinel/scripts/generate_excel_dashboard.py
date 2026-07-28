#!/usr/bin/env python3
import os
import json
import logging
import duckdb
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import re
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010\013\014\016-\037]')

def clean_string(val):
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

def format_number(val, is_currency=False):
    if is_currency:
        return f"{val:,.0f} CFA"
    return f"{val:,.0f}"

def main():
    logging.info("=== GENERATING EXECUTIVE EXCEL DASHBOARD ===")
    
    # 1. Load pre-calculated stats
    with open('dashboard_data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    kpis = data['kpis']
    geo = data['geography']
    logis = data['logistics']
    risk = data['risk_profile']
    segmentation = data['segmentation']
    top_products = data['top_products']
    forecasting = data['forecasting']
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------
    # SHEET 1: EXECUTIVE DASHBOARD
    # ---------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Executive Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10, color="333333")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="333333")
    font_kpi_val = Font(name="Segoe UI", size=14, bold=True, color="1F497D")
    font_kpi_lbl = Font(name="Segoe UI", size=9, italic=True, color="555555")
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_kpi = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_accent = PatternFill(start_color="E6EEF8", end_color="E6EEF8", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_kpi = Border(left=Side(style='medium', color="1F497D"), right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # A. Header Title
    ws_dash.merge_cells("A1:H1")
    ws_dash["A1"] = "TABLEAU DE BORD EXÉCUTIF - DOUANE SÉNÉGAL (GAINDE 2000 / ORBUS)"
    ws_dash["A1"].font = font_title
    ws_dash["A1"].alignment = align_center
    ws_dash.row_dimensions[1].height = 40
    
    # B. KPI Cards Section
    ws_dash["A3"] = "INDICATEURS CLÉS D'ACTIVITÉ (KPI)"
    ws_dash["A3"].font = font_section
    ws_dash.row_dimensions[3].height = 20
    
    # KPI 1: Dossiers
    ws_dash.merge_cells("A4:B4")
    ws_dash.merge_cells("A5:B5")
    ws_dash["A4"] = "Total Dossiers"
    ws_dash["A4"].font = font_kpi_lbl
    ws_dash["A4"].alignment = align_center
    ws_dash["A4"].fill = fill_kpi
    ws_dash["A5"] = kpis['total_dossiers']
    ws_dash["A5"].font = font_kpi_val
    ws_dash["A5"].alignment = align_center
    ws_dash["A5"].fill = fill_kpi
    ws_dash["A5"].number_format = '#,##0'
    for r in [4, 5]:
        ws_dash[f"A{r}"].border = border_kpi
        ws_dash[f"B{r}"].border = border_kpi
        
    # KPI 2: Factures
    ws_dash.merge_cells("C4:D4")
    ws_dash.merge_cells("C5:D5")
    ws_dash["C4"] = "Total Factures"
    ws_dash["C4"].font = font_kpi_lbl
    ws_dash["C4"].alignment = align_center
    ws_dash["C4"].fill = fill_kpi
    ws_dash["C5"] = kpis['total_factures']
    ws_dash["C5"].font = font_kpi_val
    ws_dash["C5"].alignment = align_center
    ws_dash["C5"].fill = fill_kpi
    ws_dash["C5"].number_format = '#,##0'
    for r in [4, 5]:
        ws_dash[f"C{r}"].border = border_kpi
        ws_dash[f"D{r}"].border = border_kpi
        
    # KPI 3: Articles
    ws_dash.merge_cells("E4:F4")
    ws_dash.merge_cells("E5:F5")
    ws_dash["E4"] = "Total Articles"
    ws_dash["E4"].font = font_kpi_lbl
    ws_dash["E4"].alignment = align_center
    ws_dash["E4"].fill = fill_kpi
    ws_dash["E5"] = kpis['total_articles']
    ws_dash["E5"].font = font_kpi_val
    ws_dash["E5"].alignment = align_center
    ws_dash["E5"].fill = fill_kpi
    ws_dash["E5"].number_format = '#,##0'
    for r in [4, 5]:
        ws_dash[f"E{r}"].border = border_kpi
        ws_dash[f"F{r}"].border = border_kpi
        
    # KPI 4: Valeur Globale
    ws_dash.merge_cells("G4:H4")
    ws_dash.merge_cells("G5:H5")
    ws_dash["G4"] = "Valeur Globale (CFA)"
    ws_dash["G4"].font = font_kpi_lbl
    ws_dash["G4"].alignment = align_center
    ws_dash["G4"].fill = fill_kpi
    ws_dash["G5"] = kpis['total_val_cfa']
    ws_dash["G5"].font = font_kpi_val
    ws_dash["G5"].alignment = align_center
    ws_dash["G5"].fill = fill_kpi
    ws_dash["G5"].number_format = '#,##0" CFA"'
    for r in [4, 5]:
        ws_dash[f"G{r}"].border = border_kpi
        ws_dash[f"H{r}"].border = border_kpi
        
    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 25
    
    # C. Financial Performance & Risques
    ws_dash["A7"] = "PERFORMANCE FINANCIÈRE MOYENNE"
    ws_dash["A7"].font = font_section
    
    headers_perf = ["Niveau Logique", "Valeur Moyenne (CFA)"]
    for c_idx, h in enumerate(headers_perf, start=1):
        cell = ws_dash.cell(row=8, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    perf_rows = [
        ["Par Dossier", kpis['avg_val_dossier']],
        ["Par Facture", kpis['avg_val_facture']],
        ["Par Article", kpis['avg_val_article']]
    ]
    for r_idx, r in enumerate(perf_rows, start=9):
        for c_idx, val in enumerate(r, start=1):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if c_idx == 1:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0" CFA"'
                
    # D. Share of Regions (Geographical)
    ws_dash["D7"] = "PART DES ÉCHANGES PAR RÉGION"
    ws_dash["D7"].font = font_section
    
    headers_geo = ["Région Économique", "Valeur Totale (CFA)", "Part (%)"]
    for c_idx, h in enumerate(headers_geo, start=4):
        cell = ws_dash.cell(row=8, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    geo_rows = []
    for region, val in geo['region_val_split'].items():
        share = geo['region_shares'][region]
        geo_rows.append([region, val, share / 100])
        
    for r_idx, r in enumerate(geo_rows, start=9):
        for c_idx, val in enumerate(r, start=4):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if c_idx == 4:
                cell.alignment = align_left
            elif c_idx == 5:
                cell.alignment = align_right
                cell.number_format = '#,##0" CFA"'
            else:
                cell.alignment = align_right
                cell.number_format = '0.0%'
                
    # E. Logistics Transport Split
    ws_dash["A13"] = "RÉPARTITION PAR MODE DE TRANSPORT"
    ws_dash["A13"].font = font_section
    
    headers_log = ["Mode de Transport", "Valeur Moyenne (CFA)", "Quantité Moyenne", "Valeur Cumulée (CFA)"]
    for c_idx, h in enumerate(headers_log, start=1):
        cell = ws_dash.cell(row=14, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    log_rows = []
    for row in logis['transport_stats']:
        log_rows.append([row['MODE_TRANSPORT'], row['valeur_moyenne'], row['quantite_moyenne'], row['total_valeur']])
        
    for r_idx, r in enumerate(log_rows, start=15):
        for c_idx, val in enumerate(r, start=1):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if c_idx == 1:
                cell.alignment = align_left
            elif c_idx in [2, 4]:
                cell.alignment = align_right
                cell.number_format = '#,##0" CFA"'
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0.0'
                
    # F. Risk Profile
    ws_dash["F13"] = "RÉPARTITION DU RISQUE DOUANIER"
    ws_dash["F13"].font = font_section
    
    headers_risk = ["Profil de Risque", "Nombre de Dossiers", "Part (%)"]
    for c_idx, h in enumerate(headers_risk, start=6):
        cell = ws_dash.cell(row=14, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    total_dossiers_sum = sum(risk.values())
    risk_rows = [
        ["Faible risque", risk['low_risk'], risk['low_risk'] / total_dossiers_sum],
        ["Moyen risque", risk['med_risk'], risk['med_risk'] / total_dossiers_sum],
        ["Haut risque", risk['high_risk'], risk['high_risk'] / total_dossiers_sum]
    ]
    for r_idx, r in enumerate(risk_rows, start=15):
        for c_idx, val in enumerate(r, start=6):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if c_idx == 6:
                cell.alignment = align_left
                if val == "Haut risque":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif c_idx == 7:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            else:
                cell.alignment = align_right
                cell.number_format = '0.0%'
                
    # G. Segmentation des Importateurs
    ws_dash["A22"] = "SEGMENTATION CLIENTS (K-MEANS CLUSTERING)"
    ws_dash["A22"].font = font_section
    
    headers_seg = ["Segment Importateur", "Nombre d'Importateurs", "Valeur Moyenne Importée (CFA)", "Dossiers Moyens"]
    for c_idx, h in enumerate(headers_seg, start=1):
        cell = ws_dash.cell(row=23, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    seg_rows = []
    for row in segmentation:
        seg_rows.append([row['segment'], row['count'], row['avg_value'], row['avg_dossiers']])
        
    for r_idx, r in enumerate(seg_rows, start=24):
        for c_idx, val in enumerate(r, start=1):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if c_idx == 1:
                cell.alignment = align_left
            elif c_idx == 2:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif c_idx == 3:
                cell.alignment = align_right
                cell.number_format = '#,##0" CFA"'
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0.0'
                
    # H. Top Products (Expensive & Value)
    ws_dash["F22"] = "TOP 5 DES MARCHANDISES EN VALEUR"
    ws_dash["F22"].font = font_section
    
    headers_prod = ["Code Tarifaire", "Désignation", "Valeur Importée (CFA)"]
    for c_idx, h in enumerate(headers_prod, start=6):
        cell = ws_dash.cell(row=23, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    top_p_rows = []
    for row in top_products['by_value'][:5]:
        top_p_rows.append([row['NUMEROTARIFDOUANE'], row.get('DESIGNATION', 'Inconnu'), row['VALEURCFA']])
        
    for r_idx, r in enumerate(top_p_rows, start=24):
        for c_idx, val in enumerate(r, start=6):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if c_idx == 6:
                cell.alignment = align_center
            elif c_idx == 7:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0" CFA"'
                
    # Adjust column widths dynamically for beauty
    for col in ws_dash.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ---------------------------------------------
    # OTHER SHEETS: RAW DATA PREVIEWS & STATS
    # ---------------------------------------------
    # We load top 1000 rows for Dossiers, Factures, and Articles to keep the sheet light yet functional
    logging.info("Generating Raw Data Preview sheets...")
    
    conn = duckdb.connect('gainde_douane.db')
    
    # Sheet 2: Dossiers
    ws_dos = wb.create_sheet(title="Dossiers Analyse")
    ws_dos.views.sheetView[0].showGridLines = True
    df_dos_sample = conn.execute("SELECT NUMERODOSSIERTPS, TYPE_OPERATION, DATE_CREATION, STATUT_DOSSIER, MODE_TRANSPORT, NOM_IMPORTATEUR, REGIME_DOUANIER, BANQUE, TYPE_CONTENEUR, IMPORTATEUR_SEGMENT, RISK_SCORE, RISK_CLASS FROM dossiers LIMIT 2000").df()
    
    # Write dataframe for Dossiers
    ws_dos.append(list(df_dos_sample.columns))
    for r in df_dos_sample.values.tolist():
        r_clean = [None if pd.isna(val) else clean_string(val) for val in r]
        ws_dos.append(r_clean)
    # Style header
    for cell in ws_dos[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    for row in ws_dos.iter_rows(min_row=2, max_row=len(df_dos_sample)+1):
        for cell in row:
            cell.font = font_body
            cell.border = border_all
            
    # Sheet 3: Factures
    ws_fac = wb.create_sheet(title="Factures Analyse")
    ws_fac.views.sheetView[0].showGridLines = True
    df_fac_sample = conn.execute("SELECT IDTPSFACTURE, NUMERODOSSIERTPS, DEVISE, TYPE_FACTURE, NUMERO_FACTURE, DATE_FACTURE, NOM_EXPORTATEUR, PAYS_EXPORTATEUR, VALEUR_FOB_CFA, VALEUR_TOTAL_CFA, INCOTERM, MODE_REGLEMENT FROM factures LIMIT 2000").df()
    
    ws_fac.append(list(df_fac_sample.columns))
    for r in df_fac_sample.values.tolist():
        r_clean = [None if pd.isna(val) else clean_string(val) for val in r]
        ws_fac.append(r_clean)
    for cell in ws_fac[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    for row in ws_fac.iter_rows(min_row=2, max_row=len(df_fac_sample)+1):
        for cell in row:
            cell.font = font_body
            cell.border = border_all
            
    # Sheet 4: Articles
    ws_art = wb.create_sheet(title="Articles Analyse")
    ws_art.views.sheetView[0].showGridLines = True
    df_art_sample = conn.execute("SELECT IDTPSFACTURE, NUMEROTARIFDOUANE, NUMERODOSSIERTPS, DESIGNATIONCOMMERCIALE, PAYSPROVENANCE, PAYSORIGINE, UNITEMESURE, QUANTITEMESURE, POIDSNET, VALEURCFA, ORDRE, ANOMALY_IF FROM articles LIMIT 2000").df()
    
    ws_art.append(list(df_art_sample.columns))
    for r in df_art_sample.values.tolist():
        r_clean = [None if pd.isna(val) else clean_string(val) for val in r]
        ws_art.append(r_clean)
    for cell in ws_art[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    for row in ws_art.iter_rows(min_row=2, max_row=len(df_art_sample)+1):
        for cell in row:
            cell.font = font_body
            cell.border = border_all
            
    conn.close()
    
    # Save file
    output_excel = "gainde_douane_dashboard.xlsx"
    wb.save(output_excel)
    logging.info(f"=== EXCEL DASHBOARD GENERATED SUCCESSFULLY: {output_excel} ===")

if __name__ == "__main__":
    main()
